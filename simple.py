from openpyxl import load_workbook
from employee import Employee
from datetime import date, datetime
import json
import logging
max_allowed_shifts = 6
logging.basicConfig(
    filename="shift_assignment.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)



def parse_comments(raw_comment):
    """
    Parses a raw comment string into a list of (comment, commenter) tuples.

    Parameters:
        raw_comment (str): The raw comment text from a cell.

    Returns:
        list: A list of tuples, each containing (comment, commenter).
    """
    if not raw_comment:
        return []

    
    # Split the comments into a list by the delimiter '\n----\n'
    comment_list = raw_comment.split('\n----\n')
    processed_comments = []

    for item in comment_list:
        # Further split into comment and commenter using '\n\t-' as the delimiter
        parts = item.split('\n\t-')
        if len(parts) == 2:
            comment, commenter = parts
            processed_comments.append((comment.strip(), commenter.strip()))
        else:
            # If no valid structure, add as "Unknown"
            processed_comments.append((parts[0].strip(), "Unknown"))
            

    return processed_comments




def get_table_header(row):
    """
    Determine if a row is a table header based on custom rules and return the header value if it exists.
    
    :param row: The row to inspect.
    :return: The header value (string or formatted date) if the row is a header, otherwise None.
    """
    for cell in row:
        if cell.value:
            # Check if the cell contains a string with "day" or "2024"
            if isinstance(cell.value, str) and ("day" in cell.value.lower() or "2024" in cell.value.lower()):
                return cell.value.strip()  # Return the string header
            # Check if the cell contains a date
            elif isinstance(cell.value, (date, datetime)):
                return cell.value.strftime("%Y-%m-%d")  # Return the formatted date
    return None

def is_merged_cell(cell):
    """
    Check if a given cell is part of a merged cell range.

    :param cell: An openpyxl cell object to check.
    :return: True if the cell is part of a merged range, False otherwise.
    """
    worksheet = cell.parent  # Get the parent worksheet of the cell
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

def load_and_assign_shift_xlsx(file_path, sheets_to_analyze):
    """
    Load an .xlsx file, analyze specific sheets, and assign shifts based on the last commenter.
    Updates the cell value with the last commenter's name.

    Parameters:
        file_path (str): Path to the .xlsx file.
        sheets_to_analyze (list): List of sheet names to analyze.

    Returns:
        dict: A dictionary with sheet names as keys and the last person who commented assigned to each shift.
    """
    # Initialize result dictionary
    results = {}

    try:
        # Load the workbook
        workbook = load_workbook(filename=file_path, data_only=False)
        shift_assignments = {}
        # Iterate through the specified sheets
        for sheet in sheets_to_analyze:
            if sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                logging.info(f"Starting to process sheet {sheet}")
                table_header = ""
                for index, row in enumerate(worksheet.iter_rows()):
                    table_header_temp = get_table_header(row)
                    if table_header_temp:  # Check if the row marks the start of a new table
                        table_header = table_header_temp
                        logging.info(f"Got new table header - {table_header}")
                        continue
                    
                    assign_shift_to = ""
                    first_name_cell_comment_final = ""
                    last_name_cell_comment_final = ""
                    if sheet == "Kitchen":
                        time_cell, first_name_cell, last_name_cell = row[2:5]
                    else:
                        time_cell, first_name_cell, last_name_cell = row[1:4]
                    # Skip rows that already have a value or are part of a merged range, or are table header Time-FirstName-LastName

                    if is_merged_cell(first_name_cell) or is_merged_cell(last_name_cell):
                        continue
                    if time_cell.value == "Time":
                        continue
                    if first_name_cell.value:
                        logging.info(f"{first_name_cell} Value already present so skipping this cell.")
                        continue
                    if not first_name_cell.comment:
                        continue
                   
                    # Get the raw comment
                    raw_comments = first_name_cell.comment.text
                    # Parse the comment into tuples
                    processed_comments = parse_comments(raw_comments)
                    
                    assign_shift_to_tuple = ()
                    # Assign the correct last commenter to the shift
                    for comment_item in reversed(processed_comments):
                        #check if the comment is by the same person.
                        if comment_item[1] == 'Unknown':
                            logging.warning(f"{last_name_cell} - There was a problem in resolving this comment please proceed manually.")
                            continue
                        
                        if comment_item[0].lower() in comment_item[1].lower():
                            employee_obj = shift_assignments.get(comment_item[1])
                            if sheet == "Dish":
                                if employee_obj:
                                    has_more_than_allowed_shifts = True if employee_obj.total_shift_count > max_allowed_shifts else False
                                    has_conflict = employee_obj.has_conflict(table_header, time_cell)
                                    if has_more_than_allowed_shifts:
                                            logging.info(f"{first_name_cell} - {employee_obj.name} already has {max_allowed_shifts} shifts so moving to next commentor")
                                            continue
                                        
                                    if has_conflict:
                                            logging.info(f"{first_name_cell} - There was a shift conflict for {comment_item[1]} so moving to next commentor.")
                                            continue
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                                else: 
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                            else:
                                if employee_obj:
                                    has_more_than_allowed_shifts = True if employee_obj.total_shift_count > max_allowed_shifts else False
                                    has_conflict = employee_obj.has_conflict(table_header, time_cell)
                                    has_dish_shift = employee_obj.dish_room_shift_taken
                                    if not has_dish_shift:
                                        logging.info(f"{first_name_cell} - {comment_item[1]} doesn't have dish room shift so moving to next commentor.")
                                        continue
                                    if has_more_than_allowed_shifts:
                                        logging.info(f"{first_name_cell} - {employee_obj.name} already has {max_allowed_shifts} shifts so moving to next commentor")
                                        continue
                                    if has_conflict:
                                        logging.info(f"{first_name_cell} - There was a shift conflict for {comment_item[1]} so moving to next commentor.")
                                        continue
                                    
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                                else: # if employee_obj not found for non dish shifts that means person doesn't have dish shift yet.
                                    logging.info(f"{first_name_cell} - {comment_item[1]} doesn't have dish room shift so moving to next commentor.")
                                    continue
                            
                        else: 
                            logging.info(f"{first_name_cell} - {comment_item[1]} has commented for someone else so moving on to next person.")
                    if len(assign_shift_to) > 0:    
                        first_name_cell_comment_final = assign_shift_to_tuple[0]
                    else:
                        logging.info(f"{first_name_cell} - Unassigned because no valid commentator found.")
                        first_name_cell.comment = None
                        last_name_cell.comment = None
                        continue
                    
                    #process last name cell
                    # Assuming last_name_cell and first_name_cell are defined within the context of your row processing
                    
                    if not last_name_cell.value and not last_name_cell.comment:
                        # If last_name_cell doesn't have a comment, extract last name from first_name_cell's processing
                        if assign_shift_to:
                            # Split the first_name_cell value into parts assuming "FirstName LastName" format
                            name_parts = assign_shift_to.split()
                            if len(name_parts) > 1:  # Ensure there's a last name
                                last_name = name_parts[-1]
                                last_name_cell_comment_final = last_name  
                            else:
                                last_name_cell_comment_final = "Unknown"  # Handle case where there's no last name
                    elif not last_name_cell.value and last_name_cell.comment :
                        # If there is a comment in last_name_cell, verify the commenter matches assign_shift_to
                        raw_comments = last_name_cell.comment.text
                        processed_comments = parse_comments(raw_comments)
                        
                        last_commenter_tuple = processed_comments[-1]
                        if last_commenter_tuple[1] == "Unknown":
                            logging.warning(f"{last_name_cell} - There was a problem in resolving last name comment please check manually to verify.")
                            last_name_cell_comment_final = last_commenter_tuple[0]
                        else: 
                            
                            if last_commenter_tuple[1] == assign_shift_to:
                                # Valid assignment, proceed
                                last_name_cell_comment_final = last_commenter_tuple[0]
                            else:
                                # Log or handle conflict scenario
                                logging.info(f"{first_name_cell} - {last_commenter_tuple[1]} commented, but {assign_shift_to} is assigned. Assigning last name of {assign_shift_to}")
                                name_parts = assign_shift_to.split()
                                if len(name_parts) > 1:  # Ensure there's a last name
                                    last_name = name_parts[-1]
                                    last_name_cell_comment_final = last_name 
                                # Update the cell with the last commenter's name
                    first_name_cell.value = first_name_cell_comment_final
                    first_name_cell.comment = None
                    last_name_cell.value = last_name_cell_comment_final
                    last_name_cell.comment = None
                    # Store the result for the current sheet
                
                    # TODO move this logic up
                    if assign_shift_to_tuple:
                        if assign_shift_to not in shift_assignments:
                            shift_assigned_employee = Employee(assign_shift_to)
                        else: 
                            shift_assigned_employee = shift_assignments.get(assign_shift_to)
                        shift_assigned_employee.add_shift(sheet, table_header, time_cell.value)
                        shift_assignments[assign_shift_to] = shift_assigned_employee
                        
                    # add to shift assignment object here

        results = shift_assignments
        # Save changes to the workbook
        workbook.save(filename=file_path)
        print("workout processing completed")
        logging.info("Workbook processing completed successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")
        logging.error(f"An error occurred: {e}")
        for sheet in sheets_to_analyze:
            results[sheet] = f"Error: {str(e)}"

    return results

# Example function usage
file_path = "Worcester Final week Schedule 2024.xlsx"  # Replace with your .xlsx file name
sheets_to_analyze = ["Dish", "Line", "Kitchen", "Pot Room", "Stir Fry", "Sushi","International Kitchen", "Grab & Go", "Salad Room"]  # Replace with sheet names to analyze
shift_assignments = load_and_assign_shift_xlsx(file_path, sheets_to_analyze)
output_file = "shift_assignments"
print(str(shift_assignments))