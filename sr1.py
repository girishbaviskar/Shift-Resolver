from openpyxl import load_workbook
import yaml
import json


def load_config(config_file):
    """
    Load configuration from a YAML file.
    """
    with open(config_file, 'r') as file:
        return yaml.safe_load(file)


def extract_tables_and_comments(file_path, sheets_to_process):
    """
    Extract comments and organize them by tables in each sheet.

    Parameters:
        file_path (str): Path to the Excel file.
        sheets_to_process (list): List of sheet names to process.

    Returns:
        list: A list of dictionaries containing sheet name, table context, cell address, value, and comment.
    """
    workbook = load_workbook(file_path, data_only=True)
    all_comments = []

    for sheet in sheets_to_process:
        if sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            print(f"Processing sheet: {sheet}")

            current_table_context = None
            current_table_comments = []

            for row in worksheet.iter_rows():
                # Detect table headers (adjust based on your sheet structure)
                if any(cell.value for cell in row):  # Non-empty row
                    if is_table_header(row):  # Check if the row marks the start of a new table
                        if current_table_comments:
                            # Save comments from the previous table
                            all_comments.append({
                                "Sheet": sheet,
                                "Table": current_table_context,
                                "Comments": current_table_comments
                            })
                            current_table_comments = []

                        # Start a new table context
                        current_table_context = get_table_context(row)

                # Process comments in the current table
                for cell in row:
                    if cell.comment:
                        current_table_comments.append({
                            "Cell": cell.coordinate,
                            "Value": cell.value,
                            "Comment": cell.comment.text
                        })

            # Save the last table's comments
            if current_table_comments:
                all_comments.append({
                    "Sheet": sheet,
                    "Table": current_table_context,
                    "Comments": current_table_comments
                })
        else:
            print(f"Sheet '{sheet}' not found in the workbook.")

    return all_comments


def is_table_header(row):
    """
    Determine if a row is a table header based on custom rules.
    Adjust this function to match the structure of your sheet.
    """
    return any(cell.value and isinstance(cell.value, str) and "Day" in cell.value for cell in row)


def get_table_context(row):
    """
    Extract table context (e.g., date or day) from a header row.
    Adjust this function to match the structure of your sheet.
    """
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            return cell.value
    return "Unknown Table"


def save_to_json(data, output_file):
    """
    Save data to a JSON file.
    """
    with open(output_file, 'w') as file:
        json.dump(data, file, indent=4)
    print(f"Comments saved to {output_file}")


if __name__ == "__main__":
    # Load configuration
    config = load_config("config.yaml")
    
    # Extract comments
    excel_file_path = "Worcester Final week Schedule 2024.xlsx"  # Update with your file path
    sheets = config.get("sheets_to_process", [])
    output_file = config.get("output_file", "processed_comments.json")
    
    comments_data = extract_tables_and_comments(excel_file_path, sheets)
    
    # Save comments to JSON
    save_to_json(comments_data, output_file)
