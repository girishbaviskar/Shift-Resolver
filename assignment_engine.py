from openpyxl import load_workbook
from employee import Employee
from datetime import date, datetime
import logging
first_finals_week = ['2024-12-11', '2024-12-12', '2024-12-13', '2024-12-14']
second_finals_week = ['2024-12-15', '2024-12-16', '2024-12-17', '2024-12-18', '2024-12-19']
max_allowed_shifts = 6
first_final_week_max_allowed_shifts = 3
second_final_week_max_allowed_shifts = 5



logging.basicConfig(
    filename="shift_assignment.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)



def parse_comments(raw_comment):
    """
    Parses a raw comment string into a list of (comment, commenter) tuples.

    Parameters:
        raw_comment (str): The raw comment text from a cell.

    Returns:
        list: A list of tuples, each containing (comment, commenter).
    """
    if not raw_comment:
        return []

    
    # Split the comments into a list by the delimiter '\n----\n'
    comment_list = raw_comment.split('\n----\n')
    processed_comments = []

    for item in comment_list:
        # Further split into comment and commenter using '\n\t-' as the delimiter
        parts = item.split('\n\t-')
        if len(parts) == 2:
            comment, commenter = parts
            processed_comments.append((comment.strip(), commenter.strip()))
        else:
            # If no valid structure, add as "Unknown"
            processed_comments.append((parts[0].strip(), "Unknown"))
            

    return processed_comments




def get_table_header(row):
    """
    Determine if a row is a table header based on custom rules and return the header value if it exists.
    
    :param row: The row to inspect.
    :return: The header value (string or formatted date) if the row is a header, otherwise None.
    """
    for cell in row:
        if cell.value:
            # Check if the cell contains a string with "day" or "2024"
            if isinstance(cell.value, str) and ("day" in cell.value.lower() or "2024" in cell.value.lower()):
                return cell.value.strip()  # Return the string header
            # Check if the cell contains a date
            elif isinstance(cell.value, (date, datetime)):
                return cell.value.strftime("%Y-%m-%d")  # Return the formatted date
    return None

def is_merged_cell(cell):
    """
    Check if a given cell is part of a merged cell range.

    :param cell: An openpyxl cell object to check.
    :return: True if the cell is part of a merged range, False otherwise.
    """
    worksheet = cell.parent  # Get the parent worksheet of the cell
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

def has_more_than_allowed_shifts_in_first_week(employee_obj):
    return True if employee_obj.first_week_shift_count >= first_final_week_max_allowed_shifts else False
    
def has_more_than_allowed_shifts_in_second_week(employee_obj):
    # return True if employee_obj.total_shift_count > max_allowed_shifts else False
    return True if employee_obj.second_week_shift_count >= second_final_week_max_allowed_shifts else False
   
def get_name_parts(assign_shift_to):
    name_parts = assign_shift_to.split()
    if len(name_parts) > 1:  # Ensure there's a last name
        last_name = name_parts[-1]
        first_name = ' '.join(name_parts[:-1])  # Join all parts except the last one
    else:
        first_name = name_parts[0]  # Only one name provided
        last_name = ''  # No last name available
    return first_name, last_name

def load_and_assign_shift_xlsx(file_path, sheets_to_analyze):
    """
    Load an .xlsx file, analyze specific sheets, and assign shifts based on the last commenter.
    Updates the cell value with the last commenter's name.

    Parameters:
        file_path (str): Path to the .xlsx file.
        sheets_to_analyze (list): List of sheet names to analyze.

    Returns:
        dict: A dictionary with sheet names as keys and the last person who commented assigned to each shift.
    """
    # Initialize result dictionary
    results = {}

    try:
        # Load the workbook
        workbook = load_workbook(filename=file_path, data_only=False)
        shift_assignments = {}
        # Iterate through the specified sheets
        for sheet in sheets_to_analyze:
            if sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                logging.info(f"Starting to process sheet {sheet}")
                table_header = ""
                for index, row in enumerate(worksheet.iter_rows()):
                    table_header_temp = get_table_header(row)
                    if table_header_temp:  # Check if the row marks the start of a new table
                        table_header = table_header_temp
                        logging.info(f"Got new table header - {table_header}")
                        continue
                    
                    assign_shift_to = ""
                    first_name_cell_comment_final = ""
                    last_name_cell_comment_final = ""
                    first_name = ""
                    last_name = ""
                    if sheet == "Kitchen":
                        time_cell, first_name_cell, last_name_cell = row[2:5]
                    else:
                        time_cell, first_name_cell, last_name_cell = row[1:4]
                    # Skip rows that already have a value or are part of a merged range, or are table header Time-FirstName-LastName

                    if is_merged_cell(first_name_cell) or is_merged_cell(last_name_cell):
                        continue
                    if time_cell.value == "Time":
                        continue
                    if first_name_cell.value:
                        logging.info(f"{first_name_cell} Value already present so skipping this cell.")
                        continue
                    if not first_name_cell.comment:
                        continue
                   
                    # Get the raw comment
                    raw_comments = first_name_cell.comment.text
                    # Parse the comment into tuples
                    processed_comments = parse_comments(raw_comments)
                    
                    assign_shift_to_tuple = ()
                    # Assign the correct last commenter to the shift
                    for comment_item in reversed(processed_comments):
                        #check if the comment is by the same person.
                        if comment_item[1] == 'Unknown':
                            logging.warning(f"{last_name_cell} - There was a problem in resolving this comment please proceed manually.")
                            continue
                        
                        if comment_item[0].lower() in comment_item[1].lower():
                            employee_obj = shift_assignments.get(comment_item[1])
                            if sheet == "Dish" or sheet == "Pot Room":
                                if employee_obj:
                                    if table_header in first_finals_week:
                                        if has_more_than_allowed_shifts_in_first_week(employee_obj):
                                            logging.info(f"{first_name_cell} - {employee_obj.name} already has {employee_obj.first_week_shift_count}/{first_final_week_max_allowed_shifts} shifts so moving to next commentor")
                                            continue
                                    else: 
                                        if has_more_than_allowed_shifts_in_second_week(employee_obj):
                                            logging.info(f"{first_name_cell} - {employee_obj.name} already has  {employee_obj.second_week_shift_count}/{second_final_week_max_allowed_shifts} shifts so moving to next commentor")
                                            continue
                                    has_conflict = employee_obj.has_conflict(table_header, time_cell)
                                    if has_conflict:
                                        logging.info(f"{first_name_cell} - There was a shift conflict for {comment_item[1]} so moving to next commentor.")
                                        continue
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                                else: 
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                            else:
                                if employee_obj:
                                    
                                    has_conflict = employee_obj.has_conflict(table_header, time_cell)
                                    has_dish_or_pot_shift = employee_obj.dish_or_pot_shift_taken
                                    if not has_dish_or_pot_shift:
                                        logging.info(f"{first_name_cell} - {comment_item[1]} doesn't have dish or pot room shift so moving to next commentor.")
                                        continue
                                    if table_header in first_finals_week:
                                        if has_more_than_allowed_shifts_in_first_week(employee_obj):
                                            logging.info(f"{first_name_cell} - {employee_obj.name} already has {employee_obj.first_week_shift_count}/{first_final_week_max_allowed_shifts} shifts so moving to next commentor")
                                            continue
                                    else: 
                                        if has_more_than_allowed_shifts_in_second_week(employee_obj):
                                            logging.info(f"{first_name_cell} - {employee_obj.name} already has  {employee_obj.second_week_shift_count}/{second_final_week_max_allowed_shifts} shifts so moving to next commentor")
                                            continue
                                    if has_conflict:
                                        logging.info(f"{first_name_cell} - There was a shift conflict for {comment_item[1]} so moving to next commentor.")
                                        continue
                                    
                                    assign_shift_to = comment_item[1]
                                    assign_shift_to_tuple = comment_item
                                    break
                                else: # if employee_obj not found for non dish shifts that means person doesn't have dish shift yet.
                                    logging.info(f"{first_name_cell} - {comment_item[1]} doesn't have dish room shift so moving to next commentor.")
                                    continue
                            
                        else: 
                            logging.info(f"{first_name_cell} - {comment_item[1]} has commented for someone else so moving on to next person.")
                    if len(assign_shift_to) > 0:    
                        first_name, last_name = get_name_parts(assign_shift_to)
                    else:
                        logging.info(f"{first_name_cell} - Unassigned because no valid commentator found.")
                        first_name_cell.comment = None
                        last_name_cell.comment = None
                        continue
                    
                    first_name_cell.value = first_name
                    last_name_cell.value = last_name
                    first_name_cell.comment = None
                    last_name_cell.comment = None
                    
                
                    # TODO move this logic up
                    if assign_shift_to_tuple:
                        if assign_shift_to not in shift_assignments:
                            shift_assigned_employee = Employee(assign_shift_to)
                        else: 
                            shift_assigned_employee = shift_assignments.get(assign_shift_to)
                        shift_assigned_employee.add_shift(sheet, table_header, time_cell.value)
                        shift_assignments[assign_shift_to] = shift_assigned_employee
                        
                    # add to shift assignment object here

        results = shift_assignments
        # Save changes to the workbook
        workbook.save(filename=file_path)
        print("workout processing completed")
        logging.info("Workbook processing completed successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")
        logging.error(f"An error occurred: {e}")
        for sheet in sheets_to_analyze:
            results[sheet] = f"Error: {str(e)}"

    return results

# Example function usage
file_path = "Worcester Final week Schedule 2024.xlsx"  # Replace with your .xlsx file name
sheets_to_analyze = ["Dish", "Pot Room", "Line", "Kitchen", "Stir Fry", "Sushi","International Kitchen", "Grab & Go", "Salad Room"]  # Replace with sheet names to analyze
#sheets_to_analyze = ["Dish"]
shift_assignments = load_and_assign_shift_xlsx(file_path, sheets_to_analyze)
output_file = "shift_assignments"
is_this_final_week_schedule = False

print(str(shift_assignments))