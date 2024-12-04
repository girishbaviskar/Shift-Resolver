from openpyxl import load_workbook

def load_and_assign_shift_xlsx(file_path, sheets_to_analyze):
    """
    Load an .xlsx file, analyze specific sheets, and assign shifts based on the last commenter.
    Updates the C5 cell with the last commenter's name.

    Parameters:
        file_path (str): Path to the .xlsx file.
        sheets_to_analyze (list): List of sheet names to analyze.

    Returns:
        dict: A dictionary with sheet names as keys and the last person who commented assigned to each shift.
    """
    # Initialize result dictionary
    results = {}

    try:
        # Load the workbook
        workbook = load_workbook(filename=file_path, data_only=False)

        # Iterate through the specified sheets
        for sheet in sheets_to_analyze:
            if sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        # Skip cells that already have a value or are part of a merged range
                        if cell.value or isinstance(cell, type(worksheet.merged_cells.ranges)):
                            continue

                        # Get the comment at cell C5
                        comment = cell.comment.text if cell.comment else "No comment"

                        # Split the comments into a list and further into (comment, commenter) tuples
                        if cell.comment:
                            comment_list = comment.split('\n----\n')
                            processed_comments = []
                            for item in comment_list:
                                parts = item.split('\n\t-')
                                if len(parts) == 2:
                                    comment, commenter = parts
                                    processed_comments.append((comment.strip(), commenter.strip()))
                                else:
                                    processed_comments.append((parts[0].strip(), "Unknown"))

                            # Assign the last commenter to the shift
                            last_commenter_tuple = processed_comments[-1] if processed_comments else ("No comment", "Unknown")
                            shift_assignee = last_commenter_tuple[0]  # Get only the commenter

                            # Update C5 with the last commenter's name
                            cell.value = shift_assignee
                            cell.comment = None
                        else:
                            shift_assignee = "No comment"

                        # Store the result for the current sheet
                        results[sheet] = shift_assignee
                        cell.comment = None
            else:
                results[sheet] = "Sheet not found"
        
        # Save changes to the workbook
        workbook.save(filename=file_path)

    except Exception as e:
        print(f"An error occurred: {e}")
        for sheet in sheets_to_analyze:
            results[sheet] = f"Error: {str(e)}"

    return results

# Example function usage
file_path = "Worcester Final week Schedule 2024.xlsx"  # Replace with your .xlsx file name
sheets_to_analyze = ["Line"]  # Replace with sheet names to analyze
shift_assignments = load_and_assign_shift_xlsx(file_path, sheets_to_analyze)
print(str(shift_assignments))